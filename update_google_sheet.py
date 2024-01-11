import argparse
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_formatting import *
from openpyxl.utils import get_column_letter  # Import the get_column_letter function

parser = argparse.ArgumentParser(description='Update operator performance data CSV on Google Sheet')
parser.add_argument('-w', '--worksheet', type=str,
                    help='The name of the worksheet to update with data from the CSV')
parser.add_argument('-f', '--file', type=str, default='operators.csv',
                    help='The CSV data file to use (default: operators.csv)')
parser.add_argument('-c', '--credentials', type=str, default='credentials.json',
                    help='The credentials JSON file used to access the Google Sheet')
parser.add_argument('-s', '--sheet', type=str,
                    help='The name of the Google Sheet to update')
args = parser.parse_args()

csv_file_path = args.file
worksheet_name = args.worksheet
sheet_name = args.sheet
credentials_file = args.credentials

# Authenticate with Google Sheets API using your credentials
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
gc = gspread.authorize(credentials)

# Open the Google Sheet by name
worksheet = gc.open(sheet_name).worksheet(worksheet_name)

# Clear the existing data in the Google Sheet
worksheet.clear()

# Read the contents of the CSV file
with open(csv_file_path, 'r') as csv_file:
    csv_contents = csv_file.read()

# Convert the CSV contents to a list of lists
csv_data = [line.split(',') for line in csv_contents.split('\n')]

# Update the specific Google Sheet with the CSV data using named arguments and a list of lists
worksheet.update(values=csv_data, range_name='A1', value_input_option='user_entered')

# Apply formatting to the header row
header_format = CellFormat(
    horizontalAlignment='CENTER',  # Center-align the text
    textFormat=TextFormat(bold=True),
    borders={ 'bottom':{'style': 'SOLID'} }
)
format_cell_range(worksheet, 'A1:Z1', header_format)

# Apply date formatting to header cells in the fourth column or greater
date_format = CellFormat(
    numberFormat=NumberFormat(type='DATE', pattern='yyyy-mm-dd')
)
format_cell_range(worksheet, f'D1:{get_column_letter(len(csv_data[0]))}1', date_format)

# Apply 2 decimal places formatting to cells in rows 2 or higher, in the fourth column or greater
decimal_format = CellFormat(
    numberFormat=NumberFormat(type='PERCENT', pattern='0.00%')
)
num_rows = len(csv_data)
format_cell_range(worksheet, 'D2:{get_column_letter(len(csv_data[0]))}{num_rows}', decimal_format)

print(f'CSV file "{csv_file_path}" has been successfully overwritten in Google Sheet "{sheet_name}" with formatting.')